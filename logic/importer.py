import openpyxl
from models import db, Equipo, Jugador, Inscripcion, Torneo
import os

def process_tournament_excel(file_path, torneo_id):
    """
    Procesa un archivo Excel para cargar equipos y jugadores con estadísticas heredadas.
    """
    try:
        wb = openpyxl.load_workbook(file_path)
        torneo = Torneo.query.get(torneo_id)
        if not torneo:
            return None, "Torneo no encontrado"

        summary = {
            "equipos_creados": 0,
            "jugadores_creados": 0,
            "errores": []
        }

        # 1. Procesar Equipos
        if "Equipos" in wb.sheetnames:
            ws_equipos = wb["Equipos"]
            # Saltar encabezado
            for row in ws_equipos.iter_rows(min_row=2, values_only=True):
                nombre = str(row[0]).strip() if row[0] else None
                if not nombre: continue
                
                # Buscar si ya existe en este torneo
                eq = Equipo.query.filter_by(nombre=nombre, torneo_id=torneo_id).first()
                if not eq:
                    eq = Equipo(
                        nombre=nombre,
                        torneo_id=torneo_id,
                        liga_id=torneo.liga_id,
                        puntos_legacy=int(row[2]) if len(row) > 2 and row[2] is not None else 0,
                        goles_f_legacy=int(row[3]) if len(row) > 3 and row[3] is not None else 0,
                        goles_c_legacy=int(row[4]) if len(row) > 4 and row[4] is not None else 0,
                        amarillas_legacy=int(row[5]) if len(row) > 5 and row[5] is not None else 0,
                        rojas_legacy=int(row[6]) if len(row) > 6 and row[6] is not None else 0,
                        saldo_arbitraje_legacy=float(row[7]) if len(row) > 7 and row[7] is not None else 0.0
                    )
                    db.session.add(eq)
                    db.session.flush()
                    # Crear inscripción
                    ins = Inscripcion(torneo_id=torneo_id, equipo_id=eq.id, monto_pactado_inscripcion=float(torneo.costo_inscripcion or 0))
                    db.session.add(ins)
                    summary["equipos_creados"] += 1
                else:
                    # Actualizar si ya existe
                    eq.puntos_legacy = int(row[2]) if len(row) > 2 and row[2] is not None else eq.puntos_legacy
                    eq.goles_f_legacy = int(row[3]) if len(row) > 3 and row[3] is not None else eq.goles_f_legacy
                    eq.goles_c_legacy = int(row[4]) if len(row) > 4 and row[4] is not None else eq.goles_c_legacy
                    eq.saldo_arbitraje_legacy = float(row[7]) if len(row) > 7 and row[7] is not None else eq.saldo_arbitraje_legacy

        # 2. Procesar Jugadores
        if "Jugadores" in wb.sheetnames:
            ws_jugadores = wb["Jugadores"]
            for row in ws_jugadores.iter_rows(min_row=2, values_only=True):
                eq_nombre = str(row[0]).strip() if row[0] else None
                j_nombre = str(row[1]).strip() if row[1] else None
                if not eq_nombre or not j_nombre: continue
                
                eq = Equipo.query.filter_by(nombre=eq_nombre, torneo_id=torneo_id).first()
                if not eq:
                    summary["errores"].append(f"Equipo '{eq_nombre}' no encontrado para el jugador '{j_nombre}'")
                    continue
                
                # Buscar jugador
                jug = Jugador.query.filter_by(nombre=j_nombre, equipo_id=eq.id).first()
                if not jug:
                    jug = Jugador(
                        nombre=j_nombre,
                        equipo_id=eq.id,
                        liga_id=torneo.liga_id,
                        numero=int(row[2]) if len(row) > 2 and row[2] is not None else None,
                        goles_legacy=int(row[3]) if len(row) > 3 and row[3] is not None else 0,
                        amarillas_legacy=int(row[4]) if len(row) > 4 and row[4] is not None else 0,
                        rojas_legacy=int(row[5]) if len(row) > 5 and row[5] is not None else 0
                    )
                    db.session.add(jug)
                    summary["jugadores_creados"] += 1
                else:
                    # Actualizar
                    jug.goles_legacy = int(row[3]) if len(row) > 3 and row[3] is not None else jug.goles_legacy
        
        db.session.commit()
        return summary, None

    except Exception as e:
        db.session.rollback()
        return None, str(e)

def generate_sample_excel(output_path):
    wb = openpyxl.Workbook()
    
    # Hoja Equipos
    ws1 = wb.active
    ws1.title = "Equipos"
    ws1.append(["Nombre", "Abreviatura", "Puntos Actuales", "GF", "GC", "Amarillas", "Rojas", "Saldo Arbitraje"])
    ws1.append(["Real Madrid", "RM", 15, 20, 10, 5, 1, 500.0])
    ws1.append(["FC Barcelona", "FCB", 12, 18, 12, 8, 0, 0.0])
    
    # Hoja Jugadores
    ws2 = wb.create_sheet("Jugadores")
    ws2.append(["Equipo", "Nombre Completo", "Dorsal", "Goles Acumulados", "Amarillas", "Rojas"])
    ws2.append(["Real Madrid", "Cristiano Ronaldo", 7, 10, 1, 0])
    ws2.append(["FC Barcelona", "Lionel Messi", 10, 12, 2, 0])
    
    wb.save(output_path)
    return output_path
